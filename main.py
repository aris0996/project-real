from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from collections import defaultdict
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash

from sqlalchemy import and_
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from io import BytesIO

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///keuangan.db'
# app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres.isadrgmnkdggmoqinxiq:arisdevdatabase@aws-0-us-east-1.pooler.supabase.com:6543/postgres'
app.config['SECRET_KEY'] = 'rahasia'
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

KATEGORI_PEMASUKAN = [
    'Gaji',
    'Bonus',
    'Investasi',
    'Penjualan',
    'Hadiah',
    'Lainnya'
]

KATEGORI_PENGELUARAN = [
    'Makanan & Minuman',
    'Transportasi',
    'Belanja',
    'Tagihan',
    'Hiburan',
    'Kesehatan',
    'Pendidikan',
    'Investasi',
    'Lainnya'
]

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128))
    nama_lengkap = db.Column(db.String(100))
    tanggal_daftar = db.Column(db.DateTime, default=datetime.utcnow)
    transaksi = db.relationship('Transaksi', backref='user', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

@login_manager.user_loader
def load_user(id):
    return User.query.get(int(id))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = 'remember' in request.form
        
        user = User.query.filter((User.username == username) | (User.email == username)).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            flash('Berhasil masuk!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Username/email atau password salah', 'danger')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        konfirmasi_password = request.form.get('konfirmasi_password')
        nama_lengkap = request.form.get('nama_lengkap')
        
        if password != konfirmasi_password:
            flash('Password tidak cocok', 'danger')
            return redirect(url_for('register'))
        
        if User.query.filter_by(username=username).first():
            flash('Username sudah digunakan', 'danger')
            return redirect(url_for('register'))
            
        if User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar', 'danger')
            return redirect(url_for('register'))
        
        user = User(username=username, email=email, nama_lengkap=nama_lengkap)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Pendaftaran berhasil! Silakan masuk', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Berhasil keluar', 'success')
    return redirect(url_for('login'))

class Transaksi(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tanggal = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    jenis = db.Column(db.String(20), nullable=False)
    kategori = db.Column(db.String(50), nullable=False)
    jumlah = db.Column(db.Float, nullable=False)
    keterangan = db.Column(db.String(200))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

@app.route('/')
@login_required
def index():
    try:
        # Ambil parameter filter
        filter_aktif = request.args.get('filter', 'semua')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        # Query dasar
        query = Transaksi.query.filter_by(user_id=current_user.id)

        # Terapkan filter
        if filter_aktif == 'hari':
            today = datetime.now().date()
            query = query.filter(
                and_(
                    Transaksi.tanggal >= today,
                    Transaksi.tanggal < today + timedelta(days=1)
                )
            )
        elif filter_aktif == 'minggu':
            today = datetime.now().date()
            start_of_week = today - timedelta(days=today.weekday())
            query = query.filter(
                and_(
                    Transaksi.tanggal >= start_of_week,
                    Transaksi.tanggal < start_of_week + timedelta(days=7)
                )
            )
        elif filter_aktif == 'bulan':
            today = datetime.now().date()
            start_of_month = today.replace(day=1)
            if today.month == 12:
                end_of_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                end_of_month = today.replace(month=today.month + 1, day=1)
            query = query.filter(
                and_(
                    Transaksi.tanggal >= start_of_month,
                    Transaksi.tanggal < end_of_month
                )
            )
        elif start_date and end_date:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(
                and_(
                    Transaksi.tanggal >= start,
                    Transaksi.tanggal < end
                )
            )

        # Ambil data transaksi yang sudah difilter
        transaksi = query.order_by(Transaksi.tanggal.desc()).all()
        
        # Hitung total dan persentase
        total_pemasukan = sum(t.jumlah for t in transaksi if t.jenis == 'pemasukan')
        total_pengeluaran = sum(t.jumlah for t in transaksi if t.jenis == 'pengeluaran')
        total_keseluruhan = total_pemasukan + total_pengeluaran
        
        # Hitung persentase (hindari pembagian dengan nol)
        if total_keseluruhan > 0:
            pemasukan_persentase = (total_pemasukan / total_keseluruhan) * 100
            pengeluaran_persentase = (total_pengeluaran / total_keseluruhan) * 100
        else:
            pemasukan_persentase = 0
            pengeluaran_persentase = 0
        
        # Data untuk grafik kategori
        kategori_pemasukan = defaultdict(float)
        kategori_pengeluaran = defaultdict(float)
        
        for t in transaksi:
            if t.jenis == 'pemasukan':
                kategori_pemasukan[t.kategori] += t.jumlah
            else:
                kategori_pengeluaran[t.kategori] += t.jumlah
        
        # Urutkan kategori berdasarkan jumlah
        pemasukan_labels = sorted(kategori_pemasukan.keys(), 
                                key=lambda x: kategori_pemasukan[x], 
                                reverse=True)
        pemasukan_data = [kategori_pemasukan[k] for k in pemasukan_labels]
        
        pengeluaran_labels = sorted(kategori_pengeluaran.keys(), 
                                  key=lambda x: kategori_pengeluaran[x], 
                                  reverse=True)
        pengeluaran_data = [kategori_pengeluaran[k] for k in pengeluaran_labels]
        
        # Data untuk grafik tren
        labels = []
        pemasukan_bulanan = []
        pengeluaran_bulanan = []
        
        # Generate data 6 bulan terakhir
        for i in range(5, -1, -1):
            date = datetime.now() - timedelta(days=30*i)
            month_start = date.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1, day=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
            
            month_trans = query.filter(
                and_(
                    Transaksi.tanggal >= month_start,
                    Transaksi.tanggal < month_end
                )
            ).all()
            
            labels.append(date.strftime('%b %Y'))
            pemasukan_bulanan.append(sum(t.jumlah for t in month_trans if t.jenis == 'pemasukan'))
            pengeluaran_bulanan.append(sum(t.jumlah for t in month_trans if t.jenis == 'pengeluaran'))

        return render_template('home.html',
            transaksi=transaksi,
            total_pemasukan=total_pemasukan,
            total_pengeluaran=total_pengeluaran,
            saldo=total_pemasukan - total_pengeluaran,
            pemasukan_persentase=pemasukan_persentase,
            pengeluaran_persentase=pengeluaran_persentase,
            labels=labels,
            pemasukan_bulanan=pemasukan_bulanan,
            pengeluaran_bulanan=pengeluaran_bulanan,
            pemasukan_labels=pemasukan_labels,
            pemasukan_data=pemasukan_data,
            pengeluaran_labels=pengeluaran_labels,
            pengeluaran_data=pengeluaran_data,
            filter_aktif=filter_aktif,
            start_date=start_date,
            end_date=end_date,
            colors=['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b',
                   '#858796', '#5a5c69', '#2e59d9', '#17a673', '#2c9faf']
        )

    except Exception as e:
        flash(str(e), 'error')
        return render_template('home.html',
            transaksi=[],
            total_pemasukan=0,
            total_pengeluaran=0,
            saldo=0,
            pemasukan_persentase=0,
            pengeluaran_persentase=0,
            labels=[],
            pemasukan_bulanan=[],
            pengeluaran_bulanan=[],
            pemasukan_labels=[],
            pemasukan_data=[],
            pengeluaran_labels=[],
            pengeluaran_data=[],
            filter_aktif='semua',
            start_date=None,
            end_date=None,
            colors=['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b',
                   '#858796', '#5a5c69', '#2e59d9', '#17a673', '#2c9faf']
        )

@app.route('/get_kategori/<jenis>')
def get_kategori(jenis):
    if jenis == 'pemasukan':
        return jsonify(KATEGORI_PEMASUKAN)
    else:
        return jsonify(KATEGORI_PENGELUARAN)

@app.route('/transaksi', methods=['GET', 'POST'])
@login_required
def transaksi():
    if request.method == 'POST':
        # Logic untuk menambah transaksi
        jenis = request.form.get('jenis')
        kategori = request.form.get('kategori')
        jumlah = float(request.form.get('jumlah'))
        keterangan = request.form.get('keterangan')
        
        transaksi_baru = Transaksi(
            tanggal=datetime.now(),
            jenis=jenis,
            kategori=kategori,
            jumlah=jumlah,
            keterangan=keterangan,
            user_id=current_user.id
        )
        
        db.session.add(transaksi_baru)
        db.session.commit()
        
        flash('Transaksi berhasil ditambahkan!', 'success')
        return redirect(url_for('transaksi'))

    # Ambil data untuk headline
    transaksi_list = Transaksi.query.filter_by(user_id=current_user.id).all()
    
    # Hitung total pemasukan
    total_pemasukan = sum(t.jumlah for t in transaksi_list if t.jenis == 'pemasukan')
    
    # Hitung total pengeluaran
    total_pengeluaran = sum(t.jumlah for t in transaksi_list if t.jenis == 'pengeluaran')
    
    # Hitung saldo
    saldo = total_pemasukan - total_pengeluaran

    # Ambil kategori untuk dropdown
    kategori_pemasukan = [
        'Gaji', 'Bonus', 'Investasi', 'Penjualan', 
        'Hadiah', 'Lain-lain'
    ]
    
    kategori_pengeluaran = [
        'Makanan & Minuman', 'Transportasi', 'Belanja', 
        'Tagihan', 'Hiburan', 'Kesehatan', 'Pendidikan',
        'Lain-lain'
    ]

    return render_template(
        'transaksi.html',
        transaksi=transaksi_list,
        saldo=saldo,
        total_pemasukan=total_pemasukan,
        total_pengeluaran=total_pengeluaran,
        kategori_pemasukan=kategori_pemasukan,
        kategori_pengeluaran=kategori_pengeluaran
    )

@app.route('/edit_transaksi/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_transaksi(id):
    transaksi = Transaksi.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    if request.method == 'POST':
        try:
            transaksi.jenis = request.form['jenis']
            transaksi.kategori = request.form['kategori']
            transaksi.jumlah = float(request.form['jumlah'])
            transaksi.keterangan = request.form['keterangan']
            
            db.session.commit()
            flash('Transaksi berhasil diupdate!', 'success')
            return redirect(url_for('transaksi'))
            
        except Exception as e:
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
            return redirect(url_for('transaksi'))
    
    return render_template('edit_transaksi.html', transaksi=transaksi)

@app.route('/hapus_transaksi/<int:id>')
@login_required
def hapus_transaksi(id):
    try:
        transaksi = Transaksi.query.filter_by(id=id, user_id=current_user.id).first_or_404()
        db.session.delete(transaksi)
        db.session.commit()
        flash('Transaksi berhasil dihapus!', 'success')
    except Exception as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return redirect(url_for('transaksi'))

@app.route('/unduh')
@login_required
def unduh():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Keuangan"

        # Modern Color Scheme
        colors = {
            'primary': '6366F1',      # Indigo
            'secondary': '4F46E5',    # Darker Indigo
            'success': '22C55E',      # Green
            'danger': 'EF4444',       # Red
            'warning': 'F59E0B',      # Amber
            'info': '3B82F6',         # Blue
            'light': 'F3F4F6',        # Gray-100
            'dark': '1F2937',         # Gray-800
            'white': 'FFFFFF',
            'border': 'E5E7EB'        # Gray-200
        }

        # Enhanced Styles
        title_style = {
            'font': Font(name='Segoe UI', size=20, bold=True, color=colors['dark']),
            'fill': PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                bottom=Side(style='medium', color=colors['primary'])
            )
        }

        subtitle_style = {
            'font': Font(name='Segoe UI', size=11, color=colors['dark']),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

        header_style = {
            'font': Font(name='Segoe UI', size=11, bold=True, color=colors['white']),
            'fill': PatternFill(start_color=colors['primary'], end_color=colors['secondary'], fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin', color=colors['border']),
                right=Side(style='thin', color=colors['border']),
                top=Side(style='thin', color=colors['border']),
                bottom=Side(style='thin', color=colors['border'])
            )
        }

        data_style = {
            'font': Font(name='Segoe UI', size=10),
            'alignment': Alignment(vertical='center'),
            'border': Border(
                left=Side(style='thin', color=colors['border']),
                right=Side(style='thin', color=colors['border']),
                top=Side(style='thin', color=colors['border']),
                bottom=Side(style='thin', color=colors['border'])
            )
        }

        amount_style = {
            'font': Font(name='Segoe UI', size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': Border(
                left=Side(style='thin', color=colors['border']),
                right=Side(style='thin', color=colors['border']),
                top=Side(style='thin', color=colors['border']),
                bottom=Side(style='thin', color=colors['border'])
            ),
            'number_format': '_-* #,##0_-;[Red]-* #,##0_-;_-* "-"_-;_-@_-'
        }

        summary_header_style = {
            'font': Font(name='Segoe UI', size=12, bold=True, color=colors['white']),
            'fill': PatternFill(start_color=colors['secondary'], end_color=colors['secondary'], fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin', color=colors['border']),
                right=Side(style='thin', color=colors['border']),
                top=Side(style='thin', color=colors['border']),
                bottom=Side(style='thin', color=colors['border'])
            )
        }

        # Set column widths
        column_widths = {
            'A': 6,   # No
            'B': 20,  # Tanggal
            'C': 15,  # Jenis
            'D': 25,  # Kategori
            'E': 20,  # Jumlah
            'F': 40   # Keterangan
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Title Section with Logo/Brand
        ws.merge_cells('A1:F2')
        title = ws.cell(row=1, column=1, value="LAPORAN KEUANGAN")
        for key, value in title_style.items():
            setattr(title, key, value)
        ws.row_dimensions[1].height = 50

        # Subtitle with Period
        ws.merge_cells('A3:F3')
        period = ws.cell(row=3, column=1, 
                        value=f"Periode: {datetime.now().strftime('%d %B %Y')}")
        for key, value in subtitle_style.items():
            setattr(period, key, value)

        # Add spacing
        ws.row_dimensions[4].height = 20

        # Headers
        headers = ['No', 'Tanggal', 'Jenis', 'Kategori', 'Jumlah', 'Keterangan']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)
        ws.row_dimensions[5].height = 30

        # Data
        transaksi = Transaksi.query.filter_by(user_id=current_user.id).order_by(Transaksi.tanggal.desc()).all()
        kategori_pemasukan = defaultdict(float)
        kategori_pengeluaran = defaultdict(float)
        
        for idx, t in enumerate(transaksi, 1):
            row = idx + 5
            
            # Apply zebra striping
            fill_color = colors['light'] if idx % 2 == 0 else colors['white']
            row_style = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            # No
            cell = ws.cell(row=row, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = data_style['font']
            cell.border = data_style['border']
            cell.fill = row_style

            # Tanggal
            cell = ws.cell(row=row, column=2, value=t.tanggal)
            cell.font = data_style['font']
            cell.border = data_style['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = 'dd/mm/yyyy hh:mm'
            cell.fill = row_style

            # Jenis
            cell = ws.cell(row=row, column=3, value=t.jenis.title())
            cell.font = Font(name='Segoe UI', size=10, 
                           color=colors['success'] if t.jenis == 'pemasukan' else colors['danger'])
            cell.border = data_style['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = row_style

            # Kategori
            cell = ws.cell(row=row, column=4, value=t.kategori)
            cell.font = data_style['font']
            cell.border = data_style['border']
            cell.alignment = data_style['alignment']
            cell.fill = row_style

            # Jumlah
            cell = ws.cell(row=row, column=5, value=t.jumlah)
            cell.font = Font(name='Segoe UI', size=10, 
                           color=colors['success'] if t.jenis == 'pemasukan' else colors['danger'])
            cell.border = amount_style['border']
            cell.alignment = amount_style['alignment']
            cell.number_format = amount_style['number_format']
            cell.fill = row_style

            # Keterangan
            cell = ws.cell(row=row, column=6, value=t.keterangan)
            cell.font = data_style['font']
            cell.border = data_style['border']
            cell.alignment = data_style['alignment']
            cell.fill = row_style

            # Update kategori totals
            if t.jenis == 'pemasukan':
                kategori_pemasukan[t.kategori] += t.jumlah
            else:
                kategori_pengeluaran[t.kategori] += t.jumlah

        current_row = len(transaksi) + 7

        # Summary Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        summary_title = ws.cell(row=current_row, column=1, value="RINGKASAN TRANSAKSI")
        for key, value in summary_header_style.items():
            setattr(summary_title, key, value)
        ws.row_dimensions[current_row].height = 30

        current_row += 2

        # Pemasukan Summary
        ws.cell(row=current_row, column=1, value="PEMASUKAN").font = Font(name='Segoe UI', bold=True, size=11)
        current_row += 1
        
        # Header for pemasukan summary
        ws.cell(row=current_row, column=2, value="Kategori").font = Font(name='Segoe UI', bold=True, size=10)
        ws.cell(row=current_row, column=3, value="Jumlah").font = Font(name='Segoe UI', bold=True, size=10)
        ws.cell(row=current_row, column=4, value="Persentase").font = Font(name='Segoe UI', bold=True, size=10)
        current_row += 1

        total_pemasukan = sum(kategori_pemasukan.values())
        
        for kategori, jumlah in sorted(kategori_pemasukan.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=current_row, column=2, value=kategori).font = data_style['font']
            
            cell = ws.cell(row=current_row, column=3, value=jumlah)
            cell.font = Font(name='Segoe UI', size=10, color=colors['success'])
            cell.number_format = amount_style['number_format']
            cell.alignment = Alignment(horizontal='right')
            
            percentage = (jumlah / total_pemasukan * 100) if total_pemasukan > 0 else 0
            ws.cell(row=current_row, column=4, 
                   value=f"{percentage:.1f}%").alignment = Alignment(horizontal='right')
            
            current_row += 1

        current_row += 2

        # Pengeluaran Summary
        ws.cell(row=current_row, column=1, value="PENGELUARAN").font = Font(name='Segoe UI', bold=True, size=11)
        current_row += 1
        
        # Header for pengeluaran summary
        ws.cell(row=current_row, column=2, value="Kategori").font = Font(name='Segoe UI', bold=True, size=10)
        ws.cell(row=current_row, column=3, value="Jumlah").font = Font(name='Segoe UI', bold=True, size=10)
        ws.cell(row=current_row, column=4, value="Persentase").font = Font(name='Segoe UI', bold=True, size=10)
        current_row += 1

        total_pengeluaran = sum(kategori_pengeluaran.values())
        
        for kategori, jumlah in sorted(kategori_pengeluaran.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=current_row, column=2, value=kategori).font = data_style['font']
            
            cell = ws.cell(row=current_row, column=3, value=jumlah)
            cell.font = Font(name='Segoe UI', size=10, color=colors['danger'])
            cell.number_format = amount_style['number_format']
            cell.alignment = Alignment(horizontal='right')
            
            percentage = (jumlah / total_pengeluaran * 100) if total_pengeluaran > 0 else 0
            ws.cell(row=current_row, column=4, 
                   value=f"{percentage:.1f}%").alignment = Alignment(horizontal='right')
            
            current_row += 1

        current_row += 2

        # Final Summary
        summary_cells = [
            ("Total Pemasukan", total_pemasukan, colors['success']),
            ("Total Pengeluaran", total_pengeluaran, colors['danger']),
            ("Saldo", total_pemasukan - total_pengeluaran, colors['dark'])
        ]

        for label, amount, color in summary_cells:
            ws.cell(row=current_row, column=2, value=label).font = Font(name='Segoe UI', bold=True, size=11)
            
            cell = ws.cell(row=current_row, column=3, value=amount)
            cell.font = Font(name='Segoe UI', bold=True, size=11, color=color)
            cell.number_format = amount_style['number_format']
            cell.alignment = Alignment(horizontal='right')
            
            current_row += 1

        # Footer
        current_row += 2
        ws.merge_cells(f'A{current_row}:F{current_row}')
        footer = ws.cell(row=current_row, column=1, 
                        value=f"Laporan ini dibuat pada {datetime.now().strftime('%d %B %Y %H:%M:%S')}")
        footer.font = Font(name='Segoe UI', size=9, italic=True, color=colors['dark'])
        footer.alignment = Alignment(horizontal='center')

        # Save
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'laporan_keuangan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        flash(f'Terjadi kesalahan saat mengunduh: {str(e)}', 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)